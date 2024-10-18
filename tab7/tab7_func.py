import re
import os
import gradio as gr
import codecs
from docx import Document
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from tab4 import tab4_func as t4
from tab3 import tab3_func as t3
from tab1 import tab1_func as t1
from datetime import datetime
import tempfile

def replace_special_periods(text):
    text = re.sub(r'\bDr\.', 'Dr<PERIOD>', text)
    text = re.sub(r'\bdr\.', 'dr<PERIOD>', text)
    text = re.sub(r'\.com\b', '<DOTCOM>', text)
    text = re.sub(r'\.org\b', '<DOTORG>', text)
    text = text.replace(" a.m."," a[dot]m[dot]").replace(" A.M."," A[dot]M[dot]")
    text = text.replace(" p.m."," p[dot]m[dot]").replace(" P.M."," P[dot]M[dot]")
    text = text.replace("...","[ddd]")
    text = text.replace("…","[ddd]")

    return text

def restore_special_periods(text):
    text = text.replace('Dr<PERIOD>', 'Dr.')
    text = text.replace('dr<PERIOD>', 'Dr.')
    text = text.replace('<DOTCOM>', '.com')
    text = text.replace('<DOTORG>', '.org')
    text = text.replace("[dot]",".")
    text = text.replace("[ddd]","…")
    return text



'''def split_segment(segment, start_time, end_time):
    if start_time is None or end_time is None:
        return [(segment, start_time, end_time)]
    #print(f"Segment before split: {segment}")  # デバッグ出力

    # 分割する前の段階で元のテキストのピリオドを保持
    sentences = re.split(r'(?<!\.\.\.)\. ', segment)
    num_sentences = len(sentences)
    times = [(start_time + i * (end_time - start_time) / num_sentences) for i in range(num_sentences + 1)]
    #print(f"Sentences after split: {sentences}")  # デバッグ出力

    # 最後の文以外にピリオドを再追加
    sentences = [sentence + '.' if i < num_sentences - 1 else sentence for i, sentence in enumerate(sentences)]
    #print(f"Sentences after adding periods: {sentences}")  # デバッグ出力

    return '''


'''def merge_segments(segments):
    merged_segments = []
    buffer_segment = ""
    buffer_start = None
    buffer_end = None

    for segment, start, end in segments:
        if buffer_segment:
            buffer_segment += " " + segment
            buffer_end = end
            if segment.endswith('.') or segment.endswith('?'): #?で終わっていたら、mergeさせない。
                merged_segments.append((buffer_segment, buffer_start, buffer_end))
                buffer_segment = ""
                buffer_start = None
                buffer_end = None
        else:
            if segment.endswith('.') or segment.endswith('?'):
                merged_segments.append((segment, start, end))
            else:
                buffer_segment = segment
                buffer_start = start
                buffer_end = end

    if buffer_segment:
        merged_segments.append((buffer_segment, buffer_start, buffer_end))

    #print(f"Merged Segments: {merged_segments}")  # デバッグ出力

    return merged_segments'''

'''def split_segment(segment, start_time, end_time):
    if start_time is None or end_time is None:
        return [(segment, start_time, end_time)]
    
    #print(f"Segment before split: {segment}")  # デバッグ出力

    # " word." または " word?" の形式で分割
    sentences = re.split(r'(?<=\w[.!?])\s+', segment)
    num_sentences = len(sentences)
    times = [(start_time + i * (end_time - start_time) / num_sentences) for i in range(num_sentences + 1)]
    
    #print(f"Sentences after split: {sentences}")  # デバッグ出力

    return [(sentences[i], times[i], times[i+1]) for i in range(num_sentences)]'''

def split_segment(segment, start_time, end_time): # 文字数基準の分割。
    if start_time is None or end_time is None:
        return [(segment, start_time, end_time)]

    #print(f"Segment before split: {segment}")  # デバッグ出力

    # 空白で区切って単語単位で処理
    words = segment.split()

    # 分割用リスト
    sentences = []
    current_sentence = []

    for word in words:
        current_sentence.append(word)
        # 単語がピリオド, 感嘆符, 質問符で終わる場合
        if word.endswith(('.', '!', '?')):
            # 現在の文を文字列として結合し、リストに追加
            sentences.append(' '.join(current_sentence))
            current_sentence = []

    # 最後の文があれば追加
    if current_sentence:
        sentences.append(' '.join(current_sentence))

    # 文の長さ（文字数）を考慮して、時間を配分
    total_length = sum(len(sentence) for sentence in sentences)
    if total_length == 0:
        return [(segment, start_time, end_time)]  # 文がない場合、元のセグメントをそのまま返す
    
    #print(f"Sentences after split: {sentences}")  # デバッグ出力

    # タイムスタンプを文の長さに応じて割り振る
    times = [start_time]
    for sentence in sentences:
        sentence_length = len(sentence)
        duration = (end_time - start_time) * (sentence_length / total_length)
        times.append(times[-1] + duration)

    return [(sentences[i], times[i], times[i+1]) for i in range(len(sentences))]

def merge_segments(segments):
    merged_segments = []
    buffer_segment = ""
    buffer_start = None
    buffer_end = None

    for segment, start, end in segments:
        if buffer_segment:
            buffer_segment += " " + segment
            buffer_end = end
            if segment.endswith('.') or segment.endswith('?'):
                merged_segments.append((buffer_segment, buffer_start, buffer_end))
                buffer_segment = ""
                buffer_start = None
                buffer_end = None
        else:
            if segment.endswith('.') or segment.endswith('?'):
                merged_segments.append((segment, start, end))
            else:
                buffer_segment = segment
                buffer_start = start
                buffer_end = end

    if buffer_segment:
        merged_segments.append((buffer_segment, buffer_start, buffer_end))

    #print(f"Merged Segments: {merged_segments}")  # デバッグ出力

    return merged_segments

def process_vtt(lines):
    segments = []
    start_time = None
    end_time = None
    text = ""
    header = lines[0]

    # タイムスタンプの整形を適用
    lines = t4.unify_timestamps_forlist(lines, 'vtt')

    for line in lines[1:]:
        if re.match(r'^\d+$', line.strip()):
            if text:
                text = replace_special_periods(text)
                if start_time is not None and end_time is not None:
                    segments.extend(split_segment(text.strip(), start_time, end_time))
            text = ""
        elif '-->' in line:
            times = line.strip().split(' --> ')
            start_time = convert_time_to_seconds(times[0])
            end_time = convert_time_to_seconds(times[1])
        else:
            text += line.strip() + " "

    if text:
        text = replace_special_periods(text)
        if start_time is not None and end_time is not None:
            segments.extend(split_segment(text.strip(), start_time, end_time))

    merged_segments = merge_segments(segments)

    output = [header]
    segment_number = 1
    for text, start, end in merged_segments:
        text = restore_special_periods(text)
        output.append(f"{segment_number}")
        output.append(convert_seconds_to_time(start, 'vtt') + ' --> ' + convert_seconds_to_time(end, 'vtt'))
        output.append(text)
        segment_number += 1
        output.append("")  # セグメント間の改行を保持

    return '\n'.join(output)

'''def process_srt(lines):
    segments = []
    start_time = None
    end_time = None
    text = ""
    segment_index = 0

    # タイムスタンプの整形を適用
    lines = t4.unify_timestamps_forlist(lines, 'srt')

    for line in lines:
        if re.match(r'^\d+$', line.strip()):
            if text:
                text = replace_special_periods(text)
                segments.extend(split_segment(text.strip(), start_time, end_time))
            segment_index = int(line.strip())
            text = ""
        elif '-->' in line:
            times = line.strip().split(' --> ')
            start_time = convert_time_to_seconds(times[0].replace(',', '.'))
            end_time = convert_time_to_seconds(times[1].replace(',', '.'))
        else:
            text += line.strip() + " "

    if text:
        text = replace_special_periods(text)
        segments.extend(split_segment(text.strip(), start_time, end_time))

    merged_segments = merge_segments(segments)

    output = []
    segment_number = 0
    for text, start, end in merged_segments:
        text = restore_special_periods(text)
        output.append(f"{segment_number + 1}\n{convert_seconds_to_time(start,'srt').replace('.', ',')} --> {convert_seconds_to_time(end,'srt').replace('.', ',')}\n{text}\n")
        segment_number += 1

    return '\n'.join(output)'''
def convert_seconds_to_time_srt(seconds):
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    return f"{h:02}:{m:02}:{s:06.3f}".replace('.', ',')
def process_srt(lines):
    segments = []
    start_time = None
    end_time = None
    text = ""
    segment_index = 0

    for line in lines:
        if re.match(r'^\d+$', line.strip()):
            if text:
                #print(f"Text before replacing periods: {text}")  # デバッグ出力
                text = replace_special_periods(text)
                #print(f"Text after replacing periods: {text}")  # デバッグ出力
                segments.extend(split_segment(text.strip(), start_time, end_time))
            segment_index = int(line.strip())
            text = ""
        elif '-->' in line:
            times = line.strip().split(' --> ')
            start_time = convert_time_to_seconds(times[0].replace(',', '.'))
            end_time = convert_time_to_seconds(times[1].replace(',', '.'))
        else:
            text += line.strip() + " "

    if text:
        #print(f"Text before replacing periods: {text}")  # デバッグ出力
        text = replace_special_periods(text)
        #print(f"Text after replacing periods: {text}")  # デバッグ出力
        segments.extend(split_segment(text.strip(), start_time, end_time))

    merged_segments = merge_segments(segments)

    output = []
    segment_number = 0
    for text, start, end in merged_segments:
        #print(f"Text before restoring periods: {text}")  # デバッグ出力
        text = restore_special_periods(text)
        #print(f"Text after restoring periods: {text}")  # デバッグ出力
        output.append(f"{segment_number + 1}\n{convert_seconds_to_time_srt(start)} --> {convert_seconds_to_time_srt(end)}\n{text}\n")
        segment_number += 1

    return '\n'.join(output)


def convert_time_to_seconds(time_str):
    try:
        time_parts = time_str.split(':')
        if len(time_parts) == 3:
            h, m, s = time_parts
        elif len(time_parts) == 2:
            h = 0
            m, s = time_parts
        else:
            raise ValueError(f"Unexpected time format: {time_str}")
        
        h = float(h)
        m = float(m)
        s = float(s.replace(',', '.'))
        return h * 3600 + m * 60 + s
    except ValueError as e:
        raise ValueError(f"Error converting time: {time_str}, {e}")



def convert_seconds_to_time(seconds, format_type='vtt'):
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    if format_type == 'vtt':
        return f"{h:01}:{m:02}:{s:06.3f}".replace(',', '.')
    else:
        return f"{h:02}:{m:02}:{s:06.3f}".replace('.', ',')
    
def process_file(input_file):
    if input_file is None:
        return None, None,[],pd.DataFrame({'1': [''], '2': [''], '3': ['']})

    with open(input_file, 'r',encoding='utf-8') as file:
        lines = file.readlines()

    _, file_extension = os.path.splitext(input_file)

    if file_extension.lower() == '.vtt':
        output = process_vtt(lines)
        output_file = os.path.splitext(input_file)[0] + '_edited.vtt'
    elif file_extension.lower() == '.srt':
        output = process_srt(lines)
        output_file = os.path.splitext(input_file)[0] + '_edited.srt'
    else:
        raise ValueError('Unsupported file format')
    
    timestamp_patch = datetime.now().strftime("%Y%m%d%H%M%S")
    temp_dir = os.path.join(tempfile.gettempdir(), f"tempdir_{timestamp_patch}")
    os.makedirs(temp_dir, exist_ok=True)

    temp_output_file=os.path.join(temp_dir,output_file)
    with open(temp_output_file, 'w',encoding='utf-8') as file:
        file.write(output)

    # Add the output to a .docx file
    '''doc = Document()
    doc.add_paragraph(output)
    basename = os.path.splitext(input_file)[0]
    timestamp_patch = datetime.now().strftime("%Y%m%d%H%M%S")
    temp_dir = os.path.join(tempfile.gettempdir(), f"tempdir_{timestamp_patch}")
    os.makedirs(temp_dir, exist_ok=True)
    if file_extension.lower() == '.vtt':
        docx_file = basename + '_edited_vtt.docx'
        
    elif file_extension.lower() == '.srt':
        docx_file = basename + '_edited_srt.docx'
     
    else:
        pass
    temp_docx_file=os.path.join(temp_dir,docx_file)
    doc.save(temp_docx_file)'''

    t7_excel_file,t7_df=t3.create_excel_from_srt(english_path=temp_output_file,tail="")
    output_html = f"""<head><meta charset="UTF-8"></head><body><pre style="white-space: pre-wrap; overflow-y: auto; height: 500px; word-wrap: break-word; padding: 10px; font-family: inherit; font-size: inherit;">{output}</pre></body>"""


    t7_df=t1.dataframe_to_html_table(t7_df)
    t7_df=f"""
        <div class="my-table-container">
            {t7_df}
        </div>
    """     
    return output_html, [temp_output_file,t7_excel_file],temp_output_file,t7_df

##翻訳後の関数##
def correct_srt_format_from_text(text):
    # Remove all whitespace and newlines
    content = re.sub(r'[\u200B-\u200D\uFEFF]', '', text)
    content = content.replace('\u200B', '')
    if "-->" in content:
        #content = re.sub(r'\s+', '', content)
        # Reconstruct the SRT format
        content=content.replace("\n","")
        content=content.replace(" ","^").replace("^^","^")
        content = re.sub(r'\s+', '', content) 

        patterns_replacements = [
            (r'(\d)\^+(\d)', r'\1\2'),  # ① 数字と数字の間
            (r'(\d)\^+(,)', r'\1\2'),   # ④ 数字とカンマの間
            (r'(,)\^+(\d)', r'\1\2'),   # ⑤ カンマと数字の間
            (r'(\d)\^+(:)', r'\1\2'),   # ⑥ 数字とコロンの間
            (r'(:)\^+(\d)', r'\1\2'),   # ⑦ コロンと数字の間
        ]    
        for eachpattern, replacement in patterns_replacements:
            content = re.sub(eachpattern, replacement, content)    
        

        pattern = re.compile(r'(\d{1,4})\^*(\d{2}:\d{2}:\d{2},\d{3}\^*-->\^*\d{2}:\d{2}:\d{2},\d{3})')
        #pattern = re.compile(r'(\d{1,4})\s*(\d{2}:\d{2}:\d{2},\d{3})\s*-->\s*(\d{2}:\d{2}:\d{2},\d{3})')
        matches = pattern.findall(content)
        
        segments = pattern.split(content)
        corrected_content = []
        
        for i in range(1, len(segments), 3):
            segment_id = segments[i].replace("^","")
            timestamp = segments[i + 1].replace("^","")
            text = segments[i + 2].replace("^"," ")
            
            corrected_content.append(f"{segment_id}")
            corrected_content.append(timestamp.replace('-->', ' --> '))
            corrected_content.append(text)


    else:
        # Normalize line endings
        content = re.sub(r'\r\n', '\n', content)
        # Remove multiple empty lines
        content = re.sub(r'\n\n+', '\n', content)
        # Normalize spaces
        content = re.sub(r'\t+', '\t', content)

        #print(f"After normalization: {repr(content)}")  # デバッグ用に表示

        # Split content into lines and process each line
        content = content.strip().split('\n')
        corrected_content = []

        for line in content:
            parts = line.split('\t')
            if len(parts) < 4:
                continue  # 不正な行をスキップ
            index = parts[0].strip()
            start_time = parts[1].strip()
            end_time = parts[2].strip()
            text = parts[3].strip()
            
            corrected_content.append(index)
            corrected_content.append(f"{start_time} --> {end_time}")
            corrected_content.append(text)
            #corrected_content.append('')
            
        # Remove any trailing empty elements
        while corrected_content and corrected_content[-1] == '':
            corrected_content.pop()

    # Ensure each subtitle block is separated by exactly one empty line
    final_content = "WEBVTT\n\n" + "\n\n".join("\n".join(block) for block in zip(*[iter(corrected_content)]*3))
    #final_content = "\n\n".join("\n".join(corrected_content[i:i+4]) for i in range(0, len(corrected_content), 4))
    return final_content


def correct_vtt_format_from_text(text):
    # Remove all whitespace and newlines
    content = re.sub(r'[\u200B-\u200D\uFEFF]', '', text)
    content = content.replace("\u200B","")
    # Remove all whitespace and newlines


    if "-->" in content: 
        # Reconstruct the VTT format
        content =webvtt_remover(content)
        #content = re.sub(r'\s+', '', content)
        content=content.replace("\n","")
        content=content.replace(" ","^").replace("^^","^")
        content = re.sub(r'\s+', '', content) 

        patterns_replacements = [
            (r'(\d)\^+(\d)', r'\1\2'),  # ① 数字と数字の間
            (r'(\d)\^+(,)', r'\1\2'),   # ④ 数字とカンマの間
            (r'(,)\^+(\d)', r'\1\2'),   # ⑤ カンマと数字の間
            (r'(\d)\^+(:)', r'\1\2'),   # ⑥ 数字とコロンの間
            (r'(:)\^+(\d)', r'\1\2'),   # ⑦ コロンと数字の間
        ]    
        for eachpattern, replacement in patterns_replacements:
            content = re.sub(eachpattern, replacement, content)    
        

        pattern = re.compile(r'(\d{1,4})\^*(\d{1}:\d{2}:\d{2}\.\d{3}\^*-->\^*\d{1}:\d{2}:\d{2}\.\d{3})')
        #pattern = re.compile(r'(\d{1,4})(\d{1}:\d{2}:\d{2}\.\d{3})\s*-->\s*(\d{1}:\d{2}:\d{2}\.\d{3})')
        matches = pattern.findall(content)
        #print(f"★{matches}")
        if len(matches)==0:
            pattern = re.compile(r'(\d{1,4})\^*(\d{2}:\d{2}:\d{2}\.\d{3}\^*-->\^*\d{1}:\d{2}:\d{2}\.\d{3})')
            matches = pattern.findall(content)
            #print(matches)

        #print(f'content:{content}')       
        segments = pattern.split(content)
        #print(f'segments:{segments}')
        corrected_content = []
        
        for i in range(1, len(segments), 3):
            segment_id = segments[i].replace("^","")
            timestamp = segments[i + 1].replace("^","")
            text = segments[i + 2].replace("^"," ")
            
            corrected_content.append(f"{segment_id}")
            corrected_content.append(timestamp.replace('-->', ' --> '))
            corrected_content.append(text)
            #print(f"★{corrected_content}")


    else:

        # Normalize line endings
        content = re.sub(r'\r\n', '\n', content)
        # Remove multiple empty lines
        content = re.sub(r'\n\n+', '\n', content)
        # Normalize spaces
        content = re.sub(r'\t+', '\t', content)

        #print(f"After normalization: {repr(content)}")  # デバッグ用に表示

        # Split content into lines and process each line
        content = content.strip().split('\n')
        corrected_content = []

        for line in content:
            parts = line.split('\t')
            if len(parts) < 4:
                continue  # 不正な行をスキップ
            index = parts[0].strip()
            start_time = parts[1].strip()
            end_time = parts[2].strip()
            text = parts[3].strip()
            
            corrected_content.append(index)
            corrected_content.append(f"{start_time} --> {end_time}")
            corrected_content.append(text)
            #corrected_content.append('')

        # Remove any trailing empty elements
        while corrected_content and corrected_content[-1] == '':
            corrected_content.pop()

    # Ensure each subtitle block is separated by exactly one empty line
    #final_content = "WEBVTT\n\n" + "\n\n".join("\n".join(corrected_content[i:i+4]) for i in range(0, len(corrected_content), 4))
    final_content = "WEBVTT\n\n" + "\n\n".join("\n".join(block) for block in zip(*[iter(corrected_content)]*3))
    return final_content

def vtt_translate(input_file, translated_content,output_file):
    if input_file==None or translated_content==None or output_file==None:
        return None
    
    ja_file_name, file_extension = os.path.splitext(input_file)
    output_ja_file_path = ja_file_name + "_edited_ja" + file_extension

    timestamp_patch = datetime.now().strftime("%Y%m%d%H%M%S")
    temp_dir = os.path.join(tempfile.gettempdir(), f"tempdir_{timestamp_patch}")
    os.makedirs(temp_dir, exist_ok=True)

    temp_output_ja_file_path=os.path.join(temp_dir,output_ja_file_path)

    if file_extension==".srt":
        corrected_content=correct_srt_format_from_text(translated_content)
    
    elif file_extension==".vtt":
        corrected_content=correct_vtt_format_from_text(translated_content)
    
    with open(temp_output_ja_file_path,'w',encoding='utf-8') as file:
            file.write(corrected_content+'\n')


    # excel出力
    output_excel_file = create_excel(output_file, temp_output_ja_file_path)
    return [output_ja_file_path,output_excel_file]

##追加ぶん
def webvtt_remover(sentence): #tab3用。
    # 特殊文字を削除
    sentence = re.sub(r'[\u200B-\u200D\uFEFF]', '', sentence)
    sentence = sentence.replace("\u200B","")
    # 改行文字を統一
    sentence = sentence.replace('\r\n', '\n').replace('\r', '\n')
    
    # VTTフォーマットを修正する正規表現パターン
    pattern = re.compile(r'(\d+)\n(\d{1}:\d{2}:\d{2}\.\d{3} --> \d{1}:\d{2}:\d{2}\.\d{3})', re.DOTALL)

    # 最初の一致部分を検索
    match = pattern.search(sentence)

    # 一致部分の開始インデックスを取得
    if match:
        start_index = match.start(1)  # 第1キャプチャグループの開始位置を取得
        rm_webvtt_sentence = sentence[start_index:]
    elif re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2}\.\d{3} --> \d{2}:\d{2}:\d{2}\.\d{3})', re.DOTALL).search(sentence):
        pattern = re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2}\.\d{3} --> \d{2}:\d{2}:\d{2}\.\d{3})', re.DOTALL)
        match = pattern.search(sentence)
        start_index = match.start(1)  # 第1キャプチャグループの開始位置を取得
        rm_webvtt_sentence = sentence[start_index:]
    else:
        rm_webvtt_sentence = sentence
    
    return rm_webvtt_sentence

def webvtt_rm(dic):

    sentence="".join(dic)
    # 改行文字を統一
    sentence = sentence.replace('\r\n', '\n').replace('\r', '\n')
    
    # VTTフォーマットを修正する正規表現パターン
    pattern = re.compile(r'(\d+)\n(\d{1}:\d{2}:\d{2}\.\d{3} --> \d{1}:\d{2}:\d{2}\.\d{3})', re.DOTALL)

    # 最初の一致部分を検索
    match = pattern.search(sentence)

    # 一致部分の開始インデックスを取得
    if match:
        start_index = match.start(1)  # 第1キャプチャグループの開始位置を取得
        rm_webvtt_sentence = sentence[start_index:]
    elif re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2}\.\d{3} --> \d{2}:\d{2}:\d{2}\.\d{3})', re.DOTALL).search(sentence):
        pattern = re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2}\.\d{3} --> \d{2}:\d{2}:\d{2}\.\d{3})', re.DOTALL)
        match = pattern.search(sentence)
        start_index = match.start(1)  # 第1キャプチャグループの開始位置を取得
        rm_webvtt_sentence = sentence[start_index:]
    else:
        rm_webvtt_sentence = sentence
    removed_dic=rm_webvtt_sentence.splitlines()

    return removed_dic

def create_excel(output_file, output_ja_file_path):
    segments = []
    if output_file.lower().endswith('.vtt') or output_file.lower().endswith('.srt'):
        with open(output_file, 'r',encoding='utf-8') as f:
            lines = f.readlines()
            if output_file.lower().endswith('.vtt'):
                lines = webvtt_rm(lines)
            segments.extend(parse_segments(lines))
    
    if output_ja_file_path.lower().endswith('.vtt') or output_ja_file_path.lower().endswith('.srt'):
        with codecs.open(output_ja_file_path, 'r', 'utf-8') as f:
            ja_lines = f.readlines()
            if output_ja_file_path.lower().endswith('.vtt'):
                ja_lines = webvtt_rm(ja_lines)
            ja_segments = parse_segments(ja_lines)
    
    excel_data = []
    for (eng_segment, start, end), ja_segment in zip(segments, ja_segments):
        if start is not None and end is not None:
            excel_data.append({
                'ID': segments.index((eng_segment, start, end)) + 1,
                'Start': convert_seconds_to_time(start),
                'End': convert_seconds_to_time(end),
                'English Subtitle': eng_segment,
                'Japanese Subtitle': ja_segment[0]
            })
        else:
            print(f"Skipping segment due to missing time: {eng_segment}")

    df = pd.DataFrame(excel_data)
    if output_ja_file_path.lower().endswith('.vtt') :
        output_excel_file = os.path.splitext(output_file)[0] + '.xlsx'
    elif output_ja_file_path.lower().endswith('.srt'):
        output_excel_file = os.path.splitext(output_file)[0] + '.xlsx'
    else:
        pass
    df.to_excel(output_excel_file, index=False)

    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Subtitles')
        workbook = writer.book
        worksheet = writer.sheets['Subtitles']

        column_widths = {'A': 7, 'B': 25, 'C': 25, 'D': 90, 'E': 90}
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                if cell.column_letter == 'A':
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif cell.column_letter in ['B', 'C']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column_letter in ['D', 'E']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            worksheet.row_dimensions[row[0].row].height = 30

        header_font = Font(bold=True)
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

    return output_excel_file

def parse_segments(lines):
    segments = []
    start_time = None
    end_time = None
    text = ""
    for line in lines:
        if re.match(r'^\d+$', line.strip()):
            if text:
                segments.append((text.strip(), start_time, end_time))
            text = ""
        elif '-->' in line:
            times = line.strip().split(' --> ')
            try:
                start_time = convert_time_to_seconds(times[0])
                end_time = convert_time_to_seconds(times[1])
            except ValueError as e:
                print(f"Error parsing time: {times}, {e}")
                start_time, end_time = None, None
        else:
            text += line.strip() + " "
    if text:
        segments.append((text.strip(), start_time, end_time))
    return segments
