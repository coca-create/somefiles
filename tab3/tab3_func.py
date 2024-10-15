##Tab3 2つのSRTファイルからデータフレーム表示とExcelファイルを作成します。
import gradio as gr
import pandas as pd
import os
import tempfile
import re
from openpyxl.styles import Alignment, Font, PatternFill
from tab7 import tab7_func as t7
from tab4 import tab4_func as t4
from datetime import datetime
# SRTファイルを解析する関数
def parse_srt(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    
    # タイムスタンプの整形を適用
    lines = t4.unify_timestamps_forlist(lines, 'srt')
    
    # 再度文字列として結合
    content = ''.join(lines).replace('\u200B', '')
    pattern = re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2},\d{3}) --> (\d{2}:\d{2}:\d{2},\d{3})\n(.*?)(?=\n\d|\Z)', re.DOTALL)
    matches = pattern.findall(content)
    #print(f"parse_srt_len(matches):{len(matches)}")

    if len(matches) == 0:
        pattern = re.compile(r'(\d+)\n(\d{1}:\d{2}:\d{2}\,\d{3}) --> (\d{1}:\d{2}:\d{2}\,\d{3})\n(.*?)(?=\n\d|\Z)', re.DOTALL)
        matches = pattern.findall(content)
        #print(f"parse_vtt_d2:Len(matches): {len(matches)}")

    subtitles = []
    for match in matches:
        subtitles.append({
            'ID': int(match[0]),
            'Start': match[1],
            'End': match[2],
            'Text': match[3].replace('\n', ' ')
        })
    
    return subtitles

def parse_vtt(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    # タイムスタンプの整形を適用
    lines = t4.unify_timestamps_forlist(lines, 'vtt')

    # 再度文字列として結合
    content = ''.join(lines).replace('\u200B', '')
    content = t7.webvtt_remover(content)
    #print(f"after_webvtt_remover: {content}")
    
    pattern = re.compile(r'(\d+)\n(\d{1}:\d{2}:\d{2}\.\d{3}) --> (\d{1}:\d{2}:\d{2}\.\d{3})\n(.*?)(?=\n\d|\Z)', re.DOTALL)
    matches = pattern.findall(content)
    #print(f"parse_vtt_d1:Len(matches): {len(matches)}")
    
    if len(matches) == 0:
        pattern = re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2}\.\d{3}) --> (\d{2}:\d{2}:\d{2}\.\d{3})\n(.*?)(?=\n\d|\Z)', re.DOTALL)
        matches = pattern.findall(content)
        #print(f"parse_vtt_d2:Len(matches): {len(matches)}")
        
    subtitles = []
    for match in matches:
        subtitles.append({
            'ID': int(match[0]),
            'Start': match[1],
            'End': match[2],
            'Text': match[3].replace('\n', ' ')
        })
    
    return subtitles

# SRTファイルからExcelファイルを作成する関数
def create_excel_from_srt(english_path=None, japanese_path=None,tail=""):
    if english_path and japanese_path:
        _, file_extension_en = os.path.splitext(english_path)
        _, file_extension_ja = os.path.splitext(japanese_path)

        if file_extension_en.lower() == '.vtt' and file_extension_ja.lower()=='.vtt':
            english_subtitles = parse_vtt(english_path)
            japanese_subtitles = parse_vtt(japanese_path)
        elif file_extension_en=='.srt' and file_extension_ja=='.srt':
            english_subtitles = parse_srt(english_path)
            japanese_subtitles = parse_srt(japanese_path)
        else:
            print('ファイル形式が一致しません')
            return None,None

        data = []
        for eng, jap in zip(english_subtitles, japanese_subtitles):
            data.append({
                'ID': eng['ID'],
                'Start': eng['Start'],
                'End': eng['End'],
                'English Subtitle': eng['Text'],
                'Japanese Subtitle': jap['Text']
            })

        df = pd.DataFrame(data)
        base_name = os.path.splitext(os.path.basename(english_path))[0]
        if file_extension_en=='.srt':
            excel_file_name = f"{base_name}.xlsx"
        elif file_extension_en=='.vtt':
            excel_file_name = f"{base_name}.xlsx"

    elif english_path:
        _, file_extension_en = os.path.splitext(english_path)
        if file_extension_en.lower()=='.srt':
            english_subtitles = parse_srt(english_path)
        elif file_extension_en.lower()=='.vtt':
            english_subtitles = parse_vtt(english_path)
        else:
            print('入力に誤りがあります。')
            return None,None

        data = []
        for eng in english_subtitles:
            data.append({
                'ID': eng['ID'],
                'Start': eng['Start'],
                'End': eng['End'],
                'English Subtitle': eng['Text']
            })

        df = pd.DataFrame(data)
        base_name = os.path.splitext(os.path.basename(english_path))[0]
        if file_extension_en.lower()=='.srt':
            excel_file_name = f"{base_name}{tail}_srt.xlsx"
        elif file_extension_en.lower()=='.vtt':
            excel_file_name = f"{base_name}{tail}_vtt.xlsx"

    elif japanese_path:
        _, file_extension_ja = os.path.splitext(japanese_path)
        if file_extension_ja.lower()=='.srt':
            japanese_subtitles = parse_srt(japanese_path)
        elif file_extension_ja.lower()=='.vtt':
            japanese_subtitles = parse_vtt(japanese_path)
        else:
            print('入力に誤りがあります。')
            return None,None
        data = []
        for jap in japanese_subtitles:
            data.append({
                'ID': jap['ID'],
                'Start': jap['Start'],
                'End': jap['End'],
                'Japanese Subtitle': jap['Text']
            })
        
        

        df = pd.DataFrame(data)
        base_name = os.path.splitext(os.path.basename(japanese_path))[0]
        if file_extension_ja.lower()=='.srt':
            excel_file_name = f"{base_name}_srt.xlsx"
        elif file_extension_ja.lower()=='.vtt':
            excel_file_name = f"{base_name}_vtt.xlsx"
    else:
        return None, None

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    temp_dir = os.path.join(tempfile.gettempdir(), f"tempdir_{timestamp}")
    os.makedirs(temp_dir, exist_ok=True)
    temp_file_path = os.path.join(temp_dir, excel_file_name)
   
   
    
    with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Subtitles')
        workbook = writer.book
        worksheet = writer.sheets['Subtitles']

        column_widths = {'A': 7, 'B': 25, 'C': 25, 'D': 90, 'E': 90}
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                if cell.column_letter == 'A':
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif cell.column_letter in ['B', 'C']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column_letter in ['D', 'E']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            worksheet.row_dimensions[row[0].row].height = 30

        header_font = Font(bold=True)
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        
    '''    final_excel_file_path = os.path.join(tempfile.gettempdir(), excel_file_name)
        os.rename(temp_excel_file_path, final_excel_file_path)'''

    return temp_file_path, df

# コンポーネントの表示を更新してファイルをクリアする関数pd.DataFrame({'1': [''], '2': [''],'3': ['']})
def update_visibility_and_clear(choice):
    clear_update = (gr.update(value=None, visible=True), gr.update(value=None, visible=True))# gr.update(value=pd.DataFrame({'1': [''], '2': [''],'3': ['']}), visible=True), gr.update(value=None, visible=True))
    if choice == "only English":
        return (gr.update(value=None, visible=True), gr.update(value=None, visible=False)) + clear_update
    elif choice == "only Japanese":
        return (gr.update(value=None, visible=False), gr.update(value=None, visible=True)) + clear_update
    else:  # "English and Japanese"
        return (gr.update(value=None, visible=True), gr.update(value=None, visible=True)) + clear_update

# ファイルをクリアする関数
def clear_all_files():
    return (gr.update(value=None), gr.update(value=None), gr.update(value=pd.DataFrame({'1': [''], '2': [''],'3': ['']})), gr.update(value=None))

    
